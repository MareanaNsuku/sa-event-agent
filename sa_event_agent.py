#!/usr/bin/env python3
"""
SA Event Aggregator – Maximum Coverage (2026+)
Directly searches event platforms via DuckDuckGo site: operator.
"""

import datetime, os, time, random, re
from ddgs import DDGS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =================== CONFIGURATION ===================
SEARCH_QUERIES = {
    "Networking": [
        "Johannesburg networking event 2026",
        "Johannesburg business mixer 2026",
        "Johannesburg startup meetup 2026",
        "Johannesburg young professionals networking",
        "register Johannesburg networking 2026",
        "Johannesburg professional networking event 2026",
        "Johannesburg tech networking 2026",
        "site:meetup.com Johannesburg networking",
        "site:eventbrite.com Johannesburg networking 2026",
        "site:meetup.com Johannesburg professional networking",
        "site:eventbrite.com Johannesburg business networking 2026"
    ],
    "Engineering Design": [
        "enter mechanical design competition 2026 South Africa",
        "electrical design challenge 2026 South Africa",
        "mechatronics design contest South Africa 2026",
        "engineering design competition for students 2026 South Africa",
        "civil engineering design competition 2026 South Africa",
        "student engineering design contest 2026 South Africa",
        "engineering innovation challenge South Africa 2026",
        "site:eventbrite.com engineering design competition South Africa",
        "site:meetup.com engineering design South Africa",
        "site:hackathon.com engineering design competition South Africa"
    ],
    "Hackathons": [
        "enter hackathon 2026 South Africa",
        "register hackathon Johannesburg 2026",
        "datathon 2026 South Africa",
        "student hackathon 2026 South Africa",
        "online hackathon 2026 South Africa",
        "coding competition 2026 South Africa",
        "hackathon Cape Town 2026",
        "innovation hackathon South Africa 2026",
        "site:hackathon.com South Africa",
        "site:devpost.com hackathon South Africa 2026",
        "site:eventbrite.com hackathon South Africa 2026",
        "site:meetup.com hackathon Johannesburg 2026"
    ],
    "Coding Bootcamps": [
        "apply coding bootcamp Johannesburg 2026",
        "full stack developer bootcamp 2026 South Africa",
        "data science bootcamp 2026 Cape Town",
        "software engineering immersive course 2026 South Africa",
        "free coding bootcamp South Africa 2026",
        "web development bootcamp Johannesburg 2026",
        "learn to code bootcamp South Africa 2026",
        "site:eventbrite.com coding bootcamp South Africa 2026",
        "site:meetup.com coding bootcamp Johannesburg"
    ],
    "AI Networking": [
        "AI networking event Johannesburg 2026",
        "machine learning meetup 2026 South Africa",
        "artificial intelligence conference Johannesburg 2026",
        "AI summit 2026 Cape Town",
        "deep learning networking South Africa 2026",
        "AI and data science event South Africa 2026",
        "tech meetup AI Johannesburg 2026",
        "site:meetup.com AI Johannesburg 2026",
        "site:eventbrite.com AI conference South Africa 2026"
    ],
    "Student Design": [
        "student design competition 2026 South Africa",
        "undergraduate design challenge 2026",
        "designathon university students 2026",
        "product design competition for students 2026",
        "global design contest undergraduates 2026",
        "student innovation design competition South Africa 2026",
        "student design contest South Africa 2026",
        "site:eventbrite.com design competition South Africa 2026",
        "site:hackathon.com designathon South Africa"
    ],
    "Car & Tech Prizes": [
        "win a car competition South Africa 2026",
        "tech gadget giveaway South Africa 2026",
        "win an iPhone South Africa 2026",
        "car sweepstakes South Africa 2026",
        "competition win car South Africa 2026",
        "prize draw South Africa 2026",
        "win a laptop South Africa 2026",
        "cash prize competition South Africa 2026",
        "site:competitions.co.za win",
        "site:eventbrite.com car giveaway South Africa"
    ],
    "Online Competitions": [
        "site:zindi.africa competition 2026",
        "site:kaggle.com competition",
        "site:devpost.com competition",
        "site:hackerearth.com challenge",
        "site:codechef.com contest",
        "online competition for South Africans 2026",
        "global hackathon accepting South Africa 2026"
    ]
}

SHEET_COLOURS = {
    "Networking":               "B3D9FF",
    "Engineering Design":       "FFD9B3",
    "Hackathons":               "B3FFB3",
    "Coding Bootcamps":         "FFFFB3",
    "AI Networking":            "E6B3FF",
    "Student Design":           "FFB3D9",
    "Car & Tech Prizes":        "FFD700",
    "Online Competitions":      "87CEEB"
}

BLOCKED_DOMAINS = [
    "sanews.gov.za", "worldpopulationreview.com", "gov.za", "youthop.com",
    "facebook.com", "linkedin.com", "youtube.com", "wikipedia.org",
    "dailyinvestor.com", "smfnews.org", "ewb-international.org",
    "nf-co.re", "globalsouthopportunities.com"
]
BLOCKED_PATTERNS = [re.compile(d) for d in BLOCKED_DOMAINS]

COMPETITION_KEYWORDS = [
    "competition", "contest", "challenge", "hackathon", "bootcamp",
    "datathon", "designathon", "summit", "conference", "meetup",
    "networking", "event", "register", "apply", "enter", "entrant",
    "open for", "call for", "submission",
    "win", "prize", "giveaway", "sweepstakes", "draw", "car", "tech", "gadget",
    "iphone", "samsung", "macbook", "laptop", "tv", "cash", "voucher",
    "ongoing", "open"
]

FUTURE_YEARS = ["2026", "2027", "2028", "2029", "2030"]

MAX_RESULTS_PER_QUERY = 20
QUERY_DELAY = 1.2
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "SA_Events.xlsx")
HEADERS = ["Title", "URL", "Date Found"]

def load_existing_urls_per_category():
    existing = {cat: set() for cat in SEARCH_QUERIES}
    if not os.path.exists(EXCEL_FILE):
        return existing
    wb = load_workbook(EXCEL_FILE)
    for cat in SEARCH_QUERIES:
        if cat in wb.sheetnames:
            ws = wb[cat]
            for row in ws.iter_rows(min_row=2, values_only=True):
                url = row[1]
                if url:
                    existing[cat].add(url)
    wb.close()
    return existing

def is_blocked(url):
    for pat in BLOCKED_PATTERNS:
        if pat.search(url):
            return True
    return False

def is_valid_event(title, snippet=""):
    text = (title + " " + snippet).lower()
    if not any(kw in text for kw in COMPETITION_KEYWORDS):
        return False
    if any(year in text for year in FUTURE_YEARS):
        return True
    if "ongoing" in text or "open" in text:
        return True
    past_years = ["2025", "2024", "2023", "2022", "2021", "2020", "2019", "2018"]
    if any(y in text for y in past_years):
        return False
    return True

def fetch_new_events(existing_urls_by_cat):
    new_events = []
    seen_in_run = set()
    with DDGS() as ddgs:
        for category, queries in SEARCH_QUERIES.items():
            for query in queries:
                print(f"Searching: {category} -> '{query}'")
                for attempt in range(2):
                    try:
                        results = list(ddgs.text(query, max_results=MAX_RESULTS_PER_QUERY))
                        break
                    except Exception as e:
                        if attempt == 0:
                            time.sleep(5)
                        else:
                            print(f"Error after retry: {e}")
                            results = []
                for res in results:
                    url = res.get('href')
                    title = res.get('title', '').strip()
                    snippet = res.get('body', '')
                    if not url or url in seen_in_run or url in existing_urls_by_cat[category]:
                        continue
                    if is_blocked(url):
                        continue
                    if not is_valid_event(title, snippet):
                        continue
                    seen_in_run.add(url)
                    new_events.append({
                        "Title": title,
                        "URL": url,
                        "Category": category,
                        "Date Found": datetime.date.today().isoformat()
                    })
                time.sleep(QUERY_DELAY + random.uniform(0, 0.5))
    return new_events

def format_sheet(ws, category):
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 15
    fill_color = SHEET_COLOURS.get(category, "FFFFFF")
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(bold=True, size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.sheet_properties.tabColor = fill_color

def update_excel(new_events):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)
    for cat in SEARCH_QUERIES:
        if cat not in wb.sheetnames:
            ws = wb.create_sheet(title=cat)
            ws.append(HEADERS)
    events_by_cat = {}
    for e in new_events:
        events_by_cat.setdefault(e["Category"], []).append(e)
    for cat, events in events_by_cat.items():
        ws = wb[cat]
        for e in events:
            ws.append([e["Title"], e["URL"], e["Date Found"]])
    for cat in SEARCH_QUERIES:
        format_sheet(wb[cat], cat)
    wb.save(EXCEL_FILE)
    total = sum(len(v) for v in events_by_cat.values())
    print(f"Added {total} new events across {len(events_by_cat)} categories.")

def main():
    print("=== SA Event Agent (Ultimate Coverage) ===")
    existing = load_existing_urls_per_category()
    new_events = fetch_new_events(existing)
    if new_events:
        update_excel(new_events)
    else:
        print("No new events found. Formatting existing file...")
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            for cat in SEARCH_QUERIES:
                if cat in wb.sheetnames:
                    format_sheet(wb[cat], cat)
            wb.save(EXCEL_FILE)
    print("=== Agent finished ===")

if __name__ == "__main__":
    main()
